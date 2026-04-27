"""核心升级逻辑：直接操作 xlsm zip 内的 XML 修改单元格值。

使用字符串替换而非 XML 解析，避免破坏原始 XML 结构和命名空间。
"""

from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path
from tempfile import NamedTemporaryFile

import openpyxl


def _col_to_num(col: str) -> int:
    result = 0
    for c in col.upper():
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result


def _num_to_col(num: int) -> str:
    result = ""
    while num > 0:
        num, remainder = divmod(num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _parse_cell_ref(ref: str) -> tuple[str, int]:
    match = re.match(r"^([A-Z]+)(\d+)$", ref.upper())
    if not match:
        raise ValueError(f"Invalid cell reference: {ref}")
    return match.group(1), int(match.group(2))


def _find_sheet_xml(zf: zipfile.ZipFile, sheet_name: str) -> str | None:
    """在 xlsm zip 中找到指定 sheet 名对应的 XML 文件路径。"""
    wb_xml = zf.read("xl/workbook.xml").decode()
    # 找到 sheet name 对应的 rId
    pattern = rf'name="{re.escape(sheet_name)}"[^>]*r:id="([^"]+)"'
    match = re.search(pattern, wb_xml)
    if not match:
        # 尝试另一种属性顺序
        pattern = rf'r:id="([^"]+)"[^>]*name="{re.escape(sheet_name)}"'
        match = re.search(pattern, wb_xml)
    if not match:
        return None
    r_id = match.group(1)

    rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode()
    rel_pattern = rf'Id="{re.escape(r_id)}"[^>]*Target="([^"]+)"'
    rel_match = re.search(rel_pattern, rels_xml)
    if not rel_match:
        return None
    target = rel_match.group(1)
    return f"xl/{target}" if not target.startswith("xl/") else target


def _update_cell_str(xml: str, cell_ref: str, value: object) -> str:
    """用字符串替换方式更新单元格值。"""
    ref_upper = cell_ref.upper()

    # 构建新值的 XML 片段
    if value is None:
        return xml
    elif isinstance(value, (int, float)):
        type_attr = ' t="n"'
        val_text = str(value)
    else:
        type_attr = ' t="str"'
        val_text = str(value)

    # 尝试找到已有单元格并替换
    # 匹配 <c r="REF" ...>...<v>...</v>...</c> 或自闭合 <c r="REF" .../>
    cell_pattern = rf'(<c r="{ref_upper}"[^>]*)(/>)'
    cell_match = re.search(cell_pattern, xml)

    if cell_match:
        # 自闭合单元格 -> 改为有值的
        prefix = cell_match.group(1)
        # 移除已有的 t="..." 属性
        prefix = re.sub(r'\s+t="[^"]*"', '', prefix)
        new_cell = f'{prefix}{type_attr}><v>{val_text}</v></c>'
        return xml[:cell_match.start()] + new_cell + xml[cell_match.end():]

    # 匹配已有内容的单元格
    cell_pattern = rf'(<c r="{ref_upper}"[^>]*)(>)(.*?)(</c>)'
    cell_match = re.search(cell_pattern, xml, re.DOTALL)

    if cell_match:
        prefix = cell_match.group(1)
        # 移除已有的 t="..." 属性
        prefix = re.sub(r'\s+t="[^"]*"', '', prefix)
        # 移除旧的 <v>...</v>（含带属性的 <v xml:space="preserve">）和自闭合 <v/>
        content = cell_match.group(3)
        content = re.sub(r'<v[^>]*>[^<]*</v>', '', content)
        content = re.sub(r'<v\s*/>', '', content)
        # 移除旧的 <f>...</f>（含带属性的 <f ca="1">）和自闭合 <f/> (公式)
        content = re.sub(r'<f[^>]*>[^<]*</f>', '', content)
        content = re.sub(r'<f\s*/>', '', content)
        new_cell = f'{prefix}{type_attr}>{content}<v>{val_text}</v></c>'
        return xml[:cell_match.start()] + new_cell + xml[cell_match.end():]

    # 单元格不存在 -> 在对应行中插入
    col_letter, row_num = _parse_cell_ref(ref_upper)

    # 找到对应行
    row_pattern = rf'(<row r="{row_num}"[^>]*>)(.*?)(</row>)'
    row_match = re.search(row_pattern, xml, re.DOTALL)

    if row_match:
        # 从同行其他单元格复制样式属性 s
        row_content = row_match.group(2)
        style_attr = ""
        style_match = re.search(r'\ss="(\d+)"', row_content)
        if style_match:
            style_attr = f' s="{style_match.group(1)}"'
        new_cell = f'<c r="{ref_upper}"{style_attr}{type_attr}><v>{val_text}</v></c>'
        # 按列顺序插入新单元格
        new_col_num = _col_to_num(col_letter)
        insert_pos = 0
        for cm in re.finditer(r'<c r="([A-Z]+)\d+"', row_content):
            existing_col_num = _col_to_num(cm.group(1))
            if existing_col_num > new_col_num:
                insert_pos = cm.start()
                break
        if insert_pos == 0:
            # 没找到更大的列，追加到末尾
            new_row_content = row_content + new_cell
        else:
            new_row_content = row_content[:insert_pos] + new_cell + row_content[insert_pos:]
        new_row = row_match.group(1) + new_row_content + row_match.group(3)
        return xml[:row_match.start()] + new_row + xml[row_match.end():]

    # 行不存在 -> 在 sheetData 末尾添加行
    # 找到 </sheetData> 位置
    sd_end = xml.find("</sheetData>")
    if sd_end < 0:
        return xml
    new_row = f'<row r="{row_num}"><c r="{ref_upper}"{type_attr}><v>{val_text}</v></c></row>'
    return xml[:sd_end] + new_row + xml[sd_end:]


def _batch_update_cells(
    zf: zipfile.ZipFile, zout: zipfile.ZipFile, sheet_name: str, updates: dict[str, object]
) -> None:
    """批量更新一个 sheet 中的多个单元格。"""
    sheet_path = _find_sheet_xml(zf, sheet_name)
    if sheet_path is None:
        print(f"  警告: 找不到工作表 '{sheet_name}'")
        return

    xml = zf.read(sheet_path).decode()
    for cell_ref, value in updates.items():
        xml = _update_cell_str(xml, cell_ref, value)
    zout.writestr(sheet_path, xml)


def upgrade(xlsx_path: str | Path, xlsm_path: str | Path, output_path: str | Path) -> Path:
    """执行升级：将 xlsx 数据匹配复制到 xlsm 副本，返回输出路径。"""
    xlsx_path = Path(xlsx_path)
    xlsm_path = Path(xlsm_path)
    output_path = Path(output_path)

    shutil.copy2(xlsm_path, output_path)

    wb_src = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_dst_cached = openpyxl.load_workbook(output_path, data_only=True)

    all_updates: dict[str, dict[str, object]] = {}

    def add(sheet: str, cell: str, value: object) -> None:
        if sheet not in all_updates:
            all_updates[sheet] = {}
        if value is not None:
            all_updates[sheet][cell] = value

    _op1_总表(wb_src, add)
    _op2_bom_z_s(wb_src, add)
    _op3_材料费(wb_src, wb_dst_cached, add)
    _op4_加工费(wb_src, wb_dst_cached, add)
    _op5_x部品(wb_src, add)

    wb_src.close()
    wb_dst_cached.close()

    # 预先找到需要更新的 sheet 路径
    updated_sheet_paths: set[str] = set()
    with zipfile.ZipFile(output_path, "r") as zin:
        for sheet_name in all_updates:
            sp = _find_sheet_xml(zin, sheet_name)
            if sp:
                updated_sheet_paths.add(sp)

    with zipfile.ZipFile(output_path, "r") as zin:
        with zipfile.ZipFile(output_path.with_suffix(".tmp.xlsm"), "w", zipfile.ZIP_DEFLATED) as zout:
            for sheet_name, cells in all_updates.items():
                _batch_update_cells(zin, zout, sheet_name, cells)

            for item in zin.namelist():
                if item in updated_sheet_paths:
                    continue
                # 跳过 calcChain，Excel 会在打开时重新计算
                if item == "xl/calcChain.xml":
                    continue
                data = zin.read(item)
                # 从 Content_Types 和 workbook.xml.rels 中移除 calcChain 引用
                if item == "[Content_Types].xml":
                    data = re.sub(
                        rb'<Override[^>]*calcChain[^>]*/>\s*', b"", data
                    )
                elif item == "xl/_rels/workbook.xml.rels":
                    data = re.sub(
                        rb'<Relationship[^>]*calcChain[^>]*/>\s*', b"", data
                    )
                zout.writestr(item, data)

    tmp_path = output_path.with_suffix(".tmp.xlsm")
    shutil.move(tmp_path, output_path)
    return output_path


# --- 5 个操作 ---

def _op1_总表(wb_src, add):
    ws = wb_src["总表"]
    for row in range(4, 22):
        for col in range(1, 3):
            val = ws.cell(row, col).value
            if val is not None:
                add("总表", f"{_num_to_col(col)}{row}", val)


def _op2_bom_z_s(wb_src, add):
    ws = wb_src["BOM分解"]
    for row in range(4, 7):
        val = ws.cell(row, 26).value
        if val is not None:
            add("BOM分解", f"Z{row}", val)
    val = ws.cell(9, 19).value
    if val is not None:
        add("BOM分解", "S9", val)


def _op3_材料费(wb_src, wb_dst_cached, add):
    ws_src = wb_src["材料费"]
    ws_dst = wb_dst_cached["材料费"]
    copy_cols = [7, 9, 28, 32, 47, 50, 53, 56, 59]
    dst_index: dict[str, int] = {}
    for row in range(6, ws_dst.max_row + 1):
        b_val = ws_dst.cell(row, 2).value
        if b_val is not None:
            dst_index[str(b_val)] = row
    for row in range(6, ws_src.max_row + 1):
        b_val = ws_src.cell(row, 2).value
        if b_val is None:
            continue
        key = str(b_val)
        if key not in dst_index:
            continue
        dst_row = dst_index[key]
        for col in copy_cols:
            val = ws_src.cell(row, col).value
            if val is not None:
                add("材料费", f"{_num_to_col(col)}{dst_row}", val)


def _op4_加工费(wb_src, wb_dst_cached, add):
    ws_src = wb_src["加工费"]
    ws_dst = wb_dst_cached["加工费"]
    dst_index: dict[str, int] = {}
    for row in range(11, ws_dst.max_row + 1):
        b_val = ws_dst.cell(row, 2).value
        if b_val is not None and not str(b_val).startswith("="):
            dst_index[str(b_val)] = row
    for row in range(11, ws_src.max_row + 1):
        b_val = ws_src.cell(row, 2).value
        if b_val is None:
            continue
        key = str(b_val)
        if key not in dst_index:
            continue
        dst_row = dst_index[key]
        for col in range(10, 48):
            val = ws_src.cell(row, col).value
            if val is not None:
                add("加工费", f"{_num_to_col(col)}{dst_row}", val)


def _op5_x部品(wb_src, add):
    ws = wb_src["X部品及外购件"]
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if val is not None:
                add("X部品及外购件", f"{_num_to_col(col)}{row}", val)
